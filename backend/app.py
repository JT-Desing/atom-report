import csv
import io
import json
import os
from datetime import datetime, timezone
from pathlib import PurePosixPath

import boto3
from botocore.exceptions import ClientError

S3 = boto3.client("s3")
BUCKET = os.environ["SOURCE_BUCKET"]
PREFIX = os.environ.get("SOURCE_PREFIX", "marketing/").lstrip("/")
NORMALIZED_KEY = os.environ.get("NORMALIZED_KEY", "marketing/dashboard/atom_latest.json")
SUPPORTED = {".csv", ".json", ".jsonl", ".xlsx", ".xlsm", ".parquet"}


def _latest_source():
    paginator = S3.get_paginator("list_objects_v2")
    candidates = []
    for page in paginator.paginate(Bucket=BUCKET, Prefix=PREFIX):
        for obj in page.get("Contents", []):
            suffix = PurePosixPath(obj["Key"]).suffix.lower()
            if suffix in SUPPORTED and obj["Key"] != NORMALIZED_KEY:
                candidates.append(obj)
    return max(candidates, key=lambda item: item["LastModified"]) if candidates else None


def _read_records(key, raw):
    suffix = PurePosixPath(key).suffix.lower()
    if suffix == ".csv":
        text = raw.decode("utf-8-sig")
        sample = text[:4096]
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return list(csv.DictReader(io.StringIO(text), dialect=dialect))
    if suffix == ".json":
        payload = json.loads(raw.decode("utf-8-sig"))
        return payload if isinstance(payload, list) else payload.get("records", payload.get("data", []))
    if suffix == ".jsonl":
        return [json.loads(line) for line in raw.decode("utf-8-sig").splitlines() if line.strip()]
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook
        sheet = load_workbook(io.BytesIO(raw), read_only=True, data_only=True).active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        return [dict(zip(headers, row)) for row in rows]
    if suffix == ".parquet":
        import pyarrow.parquet as parquet
        return parquet.read_table(io.BytesIO(raw)).to_pylist()
    raise ValueError(f"Formato no soportado: {suffix}")


def _json_safe(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime,)):
        return value.isoformat()
    return str(value)


def ingest_handler(event, context):
    source = _latest_source()
    if not source:
        return {"status": "waiting", "message": "No hay archivos ATOM compatibles"}
    raw = S3.get_object(Bucket=BUCKET, Key=source["Key"])["Body"].read()
    records = [{str(k).strip(): _json_safe(v) for k, v in row.items()} for row in _read_records(source["Key"], raw)]
    payload = {
        "schema_version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": {"key": source["Key"], "last_modified": source["LastModified"].isoformat()},
        "record_count": len(records),
        "records": records,
    }
    S3.put_object(
        Bucket=BUCKET,
        Key=NORMALIZED_KEY,
        Body=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        ContentType="application/json",
        CacheControl="no-store",
        ServerSideEncryption="AES256",
    )
    return {"status": "processed", "source": source["Key"], "records": len(records)}


def api_handler(event, context):
    origin = os.environ.get("ALLOWED_ORIGIN", "https://jt-desing.github.io")
    headers = {"content-type": "application/json; charset=utf-8", "access-control-allow-origin": origin, "cache-control": "no-store"}
    try:
        body = S3.get_object(Bucket=BUCKET, Key=NORMALIZED_KEY)["Body"].read().decode("utf-8")
        return {"statusCode": 200, "headers": headers, "body": body}
    except ClientError as error:
        if error.response.get("Error", {}).get("Code") in {"NoSuchKey", "404"}:
            return {"statusCode": 404, "headers": headers, "body": json.dumps({"status": "waiting", "records": []})}
        raise
